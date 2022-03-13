from open_json import build_bills_list,read_database
from openpyxl import Workbook
 

def populate_sheet(sheet, data):
    title = list()
    aux = list()
    for produto in data:
        new_item = list()
        for k, v in produto.items():
            new_item.append(v)
            if k in title:
                continue
            else:
                title.append(k)
        aux.append(new_item.copy())
    aux.insert(0,title)

    for item in aux:
        c = 1
        for info in item:
            sheet.cell(row=(aux.index(item)+1), column=c, value=info)
            c += 1


def new_excel_file():
    path = r'C:\Users\João Pedro Barreto\Desktop\Python\Exercícios do Breno\Médio\Ex.02\database.json'
    database = read_database(path)
    bills = build_bills_list(database)
    arquivo_excel = Workbook()
    planilha1 = arquivo_excel.active
    planilha1.title = "Produtos"
    planilha2 = arquivo_excel.create_sheet("Nota")
    populate_sheet(planilha1, database['products'])
    populate_sheet(planilha2, bills)

    arquivo_excel.save(filename="database.xlsx")


def main():
    new_excel_file()


if __name__ == "__main__":
    main()
